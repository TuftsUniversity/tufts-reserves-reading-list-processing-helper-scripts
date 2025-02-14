import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import sys
sys.path.append('config/')
import secrets_local
#from tkinter.filedialog import askopenfilename
import glob
import pandas as pd
import json
import re
import requests
import urllib
import xml.etree.ElementTree as ET
import time
import os
import socket

barcode_array = []

def read_excel_file(file_path):
    df = pd.read_excel(file_path, skiprows=0, header=1, engine="openpyxl", dtype={'str'})
    df = df.fillna("")


    return df
def parse_xml_response(xml_string, json_course, format_type, row, api_bib, open_library_title, open_library_author):
    results = []

    electronic_record_array = []


    # Parse XML
    #try:



    root = ET.fromstring(xml_string)




    # Iterate over <record> elements
    for record in root.findall('.//record'):

        # Check for physical format
        #if format_type in ["", "Physical", "physical"]:

        alma_title = record.findtext(".//datafield[@tag='245']/subfield[@code='a']").replace("/", "")



        alma_title = re.sub(r'[,:;."\'&]', ' ', alma_title)
        alma_title = re.sub(r'[-]', ' ', alma_title)

        alma_title = re.sub(r'\s$', r'', alma_title)
        alma_title = re.sub(r'^(The|A|An)\s', r'', alma_title)
        print('Alma title')

        print("-" + alma_title.lower() + "-")

        # print("open library title")
        open_library_title = re.sub(r'[,:;."\'&]', ' ', open_library_title)
        open_library_title = re.sub(r'[-]', ' ', open_library_title)

        open_library_title = re.sub(r'\s$', r'', open_library_title)

        open_library_title = re.sub(r'^(The|A|An)\s', r'', open_library_title)
        print("Open library title")
        print("-" + open_library_title.lower() + "-")



        author = record.findtext(".//datafield[@tag='100']/subfield[@code='a']") or \
                 record.findtext(".//datafield[@tag='110']/subfield[@code='a']") or \
                 record.findtext(".//datafield[@tag='700']/subfield[@code='a']") or \
                 record.findtext(".//datafield[@tag='710']/subfield[@code='a']")
        author = re.sub(r'(.+),$', r'\1', author)
        # print("Alma author")
        # print("-" + author + "-")
        # print("open library author")
        open_library_author = re.sub(r'.*?([^\s]+)$', r'\1', open_library_author)
        # print("-" + open_library_author + "-")


        # Convert the XML element back to a string
        pretty_xml = ET.tostring(record, encoding="unicode")

        # Print the pretty XML
        # print(pretty_xml)



        # print("open library title")
        # print(open_library_title.lower())
        # print("alma title")
        # print(alma_title.lower())
        # print("open library author")
        # print(open_library_author)
        # print('Alma author')
        # print(author)
        if (alma_title.lower() == open_library_title.lower() and open_library_title != "" and open_library_author in author and open_library_author != ""):
            # print("true\n")
            # print("record type")
            # print(record.find(".//datafield[@tag='AVA']"))
            # print(record.find(".//datafield[@tag='AVE']"))
            # print("\n\n\n")

            ava = record.find(".//datafield[@tag='AVA']")
            if ava is not None:
                # print("AVA\n\n\n")
                phys_mms_id = ava.find("subfield[@code='0']").text

                # print("phsysical_mms_id")
                # print(phys_mms_id)
                # Simulate an API call to get item details
                item_query = f"https://api-na.hosted.exlibrisgroup.com/almaws/v1/bibs/{phys_mms_id}/holdings/ALL/items?apikey={api_bib}&format=json"
                response_items = requests.get(item_query)

                if response_items.status_code == 200:
                    items = response_items.json()

                    # print(json.dumps(items))
                    for item in items.get('item', []):
                        library = item.get('holding_data', {}).get('temp_library', {}).get('desc', '') if item.get('holding_data', {}).get('in_temp_location') else item.get('item_data', {}).get('library', {}).get('desc', '')
                        location = item.get('holding_data', {}).get('temp_location', {}).get('desc', '') if item.get('holding_data', {}).get('in_temp_location') else item.get('item_data', {}).get('location', {}).get('desc', '')

                        if item['item_data']['barcode'] + json_course.get('course', [{}])[0].get('code', '') + "-" + json_course.get('course', [{}])[0].get('section', '') in barcode_array:
                            continue
                        else:
                            barcode_array.append(item['item_data']['barcode'] + json_course.get('course', [{}])[0].get('code', '') + "-" + json_course.get('course', [{}])[0].get('section', ''))

                        author = record.findtext(".//datafield[@tag='100']/subfield[@code='a']") or \
                                 record.findtext(".//datafield[@tag='700']/subfield[@code='a']") or \
                                 record.findtext(".//datafield[@tag='710']/subfield[@code='a']")

                        # print("barcode array")
                        # print(barcode_array)
                        # print("\n\n\n\n\n")
                        results.append({
                            'Title': alma_title, #str(record.findtext(".//datafield[@tag='245']/subfield[@code='a']")).replace("None", "") +
                                    # str(record.findtext(".//datafield[@tag='245']/subfield[@code='b']")).replace("None", ""),
                            'Author': author,
                            'Publisher': record.findtext(".//datafield[@tag='264']/subfield[@code='b']"),
                            'Year': record.findtext(".//datafield[@tag='264']/subfield[@code='c']"),
                            'MMS ID': phys_mms_id,
                            'ISBN': record.findtext(".//datafield[@tag='020']/subfield[@code='a']"),
                            'Course Code': json_course.get('course', [{}])[0].get('code', ''),
                            'Course Section': json_course.get('course', [{}])[0].get('section', ''),
                            'Library': library,
                            'Location': location,
                            'Call Number': item.get('holding_data', {}).get('permanent_call_number', ''),
                            'Barcode': item.get('item_data', {}).get('barcode', ''),
                            'Description': json.dumps(item.get('item_data', {}).get('description', ''), ensure_ascii=False),
                            'Returned Format': 'Physical'
                        })

                        new_dict = {}
                        for key, value in row.items():


                            new_dict[key +  " - Input"] = value
                        old_dict = results[-1]



                        results[-1] = {**new_dict, **old_dict}

                        print(results[-1]['Course Code'] + "-" + results[-1]['Course Section'] + "-" + results[-1]['Title'] + "-" + results[-1]['MMS ID'])        # Check for electronic format

                else:
                    print("item retrieval problem")
                    results.append([{
                        'Title': f'Item API call failure for {row.get("TITLE", "")}, {item_query}',
                        'Author': f'Item API call failure for results for {row.get("AUTHOR", "")}, {item_query}',
                        'Year': row.get('Year', ''),
                        'Course Code': json_course.get('course', [{}])[0].get('code', ''),
                        'Returned Format': 'N/A'
                    }])

        # if format_type in ["", "Electronic", "electronic"]:
            ave = record.find(".//datafield[@tag='AVE']")
            if ave is not None:
                # print("AVE\n\n\n")
                e_mms_id = ave.find("subfield[@code='0']").text

                if f"{e_mms_id}Electronic" in electronic_record_array:
                    continue
                else:
                    electronic_record_array.append(f"{e_mms_id}Electronic")



                results.append({
                    'Title': alma_title,
                    'Author': author,
                    'Publisher': record.findtext(".//datafield[@tag='264']/subfield[@code='b']"),
                    'Year': record.findtext(".//datafield[@tag='264']/subfield[@code='c']"),
                    'MMS ID': e_mms_id,
                    'ISBN': record.findtext(".//datafield[@tag='020']/subfield[@code='a']"),
                    'Course Code': json_course.get('course', [{}])[0].get('code', ''),
                    'Course Section': json_course.get('course', [{}])[0].get('section', ''),
                    'Returned Format': 'Electronic'
                })


                new_dict = {}
                for key, value in row.items():


                    new_dict[key +  " - Input"] = value
                old_dict = results[-1]



                results[-1] = {**new_dict, **old_dict}

                print(results[-1]['Course Code'] + "-" + results[-1]['Course Section'] + "-" + results[-1]['Title'] + "-" + results[-1]['MMS ID'])
        else:

            results.append({
                'Title': f'No results for {row.get("TITLE", "")}',
                'Author': f'No results for {row.get("AUTHOR", "")}',
                'Year': row.get('Year', ''),
                'Course Code': json_course.get('course', [{}])[0].get('code', ''),
                'Returned Format': 'N/A'
            })

    #except:

        #
        # results = [{'Title': 'NO RESULTS',
        # 'Author': 'NO RESULTS',
        # 'Contributor': 'NO RESULTS',
        # 'Publisher': r'NO RESULTS',
        # 'Date': 'NO RESULTS',
        # 'MMS ID': 'NO RESULTS',
        # 'ISBN': 'NO RESULTS',
        # 'Version': 'NO RESULTS',
        # 'Course Code': 'NO RESULTS',
        # 'Course Section': 'NO RESULTS',
        # 'Returned Format': 'NO RESULTS',
        # 'Library': 'NO RESULTS',
        # 'Location': 'NO RESULTS',
        # 'Call Number': 'NO RESULTS',
        # 'Barcode': 'NO RESULTS',
        # 'Description': 'NO RESULTS',
        # 'Citation Type': '',
        # 'section_info': 'NO RESULTS',
        # 'Item Policy': 'NO RESULTS'
        # }]
        #
        # for key, value in row.items():
        #     if key not in results[-1]:
        #         results[-1][key + " - Input"] = value
        # print(results[-1]['Course Code'] + "-" + results[-1]['Course Section'] + "-" + results[-1]['Title'] + "-" + results[-1]['MMS ID'])



    return results

def get_open_library_results(url):

    return requests.get(url)
def search_alma_sru_api(row, api_key_courses, api_bib):
    title = row.get('TITLE', '')
    author_last = row.get('AUTHOR', '')
    format_type = row.get('Format', '')
    course_number = row.get('Course Number', '')
    course_semester = row.get('Term', '')
    isbn = str(int(row.get('EAN-13', '')))


    #oclc_results = requests.get("http://api.crossref.org/works?query.bibliographic=" + str(isbn) + "&rows=1")
    openlibrary_url = "https://openlibrary.org/api/books?bibkeys=ISBN:" + str(isbn) + "&jscmd=details&format=json"

    openlibrary_results = ""
    max_retry_range = 10  # Reduced for clarity, increase as needed

    for retry in range(max_retry_range):
        try:
            # Assuming get_open_library_results is a function that makes an API call
            openlibrary_results = get_open_library_results(openlibrary_url)
            break  # Success! Exit the loop
        except requests.exceptions.Timeout:  # Catch the correct timeout exception
            print(f"Attempt {retry + 1} failed. Retrying after waiting.")
            time.sleep((retry * 2) + 30)  # Exponential backoff
        except Exception as e:
            print(f"An error occurred: {e}")
            break  # Exit on other exceptions


    if openlibrary_results != "" and openlibrary_results.status_code == 200 and "ISBN:" + str(isbn) in openlibrary_results.json():
        openlibrary_results = openlibrary_results.json()
        #print(json.dumps(oclc_results))

        #print(oclc_results['message']['items'][0]['title'][0])
        #for item in oclc_results['message']['items']:




        #if item['type'] == 'monograph' or item['type'] == 'edited-book':
        #title = item['volumeInfo']['title']
        openlibrary_results_key = openlibrary_results["ISBN:" + str(isbn)] #.details.title
        title = openlibrary_results_key['details']['title']
        print(title)
        print(isbn)




        publisher = ""
        author_first = ""
        author = ""
        try:
            author = openlibrary_results_key['details']['authors'][0]['name']
            author_first = re.sub(r'(^[^\s]+)\s([^\s]+)', r'\1', author)
            author_last = re.sub(r'^([^\s]+)*(\s)*.*?([^\s]+)$', r'\3', author)
        except:
            try:

                publisher = openlibrary_results_key['details']['publishers'][0]

            except:
                publisher = ""



        # author_last = item['author'][0]['family']
        # author_first = item['author'][0]['given']

        #print("\n" + publisher)
        #sys.exit()


        if 'F' in course_semester:
            course_semester = course_semester.replace('F', 'Fa')

        elif 'W' in course_semester:
            course_semester = course_semester.replace('W', 'Sp')
        instructor = row.get('Instructor Last Name', '')

        query = "https://tufts.alma.exlibrisgroup.com/view/sru/01TUN_INST?version=1.2&operation=searchRetrieve&recordSchema=marcxml&alma.mms_material_type=BK"

        if title and title != "":

            title = re.sub(r'[,:;."\'&]', ' ', title)
            title = re.sub(r'[-]', ' ', title)

            query += f"&query=alma.title=%22{urllib.parse.quote_plus(title)}%22"
        else:

            return [{
                'Title': f'No results for {title}',
                'Author': f'No results for {author_last}',
                'Publisher': row.get('Publisher', ''),
                'Year': row.get('Year', ''),
                'Course Code': course_number,
                'Returned Format': 'N/A'
            }]

        if author_last != "":

            #author_first = row.get('Author First', '')
            if author_first != author and author_first != "":
                query += f"%20AND%20alma.creator=%22*{urllib.parse.quote_plus(author_last + ',' + author_first)}*%22"
            else:
                query += f"%20AND%20alma.creator=%22*{urllib.parse.quote_plus(author_last)}*%22"
        # elif publisher != "":
        #     query += f"%20AND%20alma.publisher=%22*{requests.utils.quote(publisher)}*%22"
        else:

            return [{
                'Title': f'Not enough metadata for results for {title}',
                'Author': f'Not enough metadata  for {author_last}',
                #'Publisher': f'Not enough metadata  for {Publisher}',
                'Course Code': f'Not enough metadata for {course_number}',
                'Returned Format': 'N/A'
            }]

        print(query)
        response = requests.get(query)

        if response.status_code == 200:
            # Removing namespaces

            xml_string = re.sub(r'xmlns[^=]*="[^"]*"', '', response.text)
            xml_string = re.sub(r'[a-zA-Z]+:([a-zA-Z]+[=>])', r'\1', xml_string)
            print(xml_string)
            # Simulate the course API response


            course_url = f"https://api-na.hosted.exlibrisgroup.com/almaws/v1/courses?apikey={api_key_courses}&q=name~{course_semester}-{course_number}%20AND%20instructors~{instructor}&format=json"
            course_url_without_instructor = f"https://api-na.hosted.exlibrisgroup.com/almaws/v1/courses?apikey={api_key_courses}&q=name~{course_semester}-{course_number}&format=json"

            response_course = ""
            if instructor != "":

                response_course = requests.get(course_url)
            else:
                response_course = requests.get(course_url_without_instructor)



            json_course = response_course.json() if response_course.status_code == 200 else {}



            # Parse the XML response
            return parse_xml_response(xml_string, json_course, format_type, row, api_bib, title, author)
        else:
            print(response.status_code)
            print(response.text)
            return [{
                'Title': f'No Alma results for {title}',
                'Author': f'No Alma results for {author_last}',
                'Publisher': f'No Alma results for {Publisher}',
                'Course Code': f'No Alma results for {course_number}',
                'Returned Format': 'N/A'
            }]
    # else:
    #     print(json.dumps(oclc_results))
    else:

        print(openlibrary_results.status_code)
        print(openlibrary_results.text)
        results = {
            'Title': f'No Open Library result for {row.get("TITLE", "")}',
            'Author': f'No Open Library  result for  {row.get("AUTHOR", "")}',
            'Contributor': f'No Open Library  result for {row.get("Contributor", "")}',
            'Year': f'No Open Library  result for {row.get("Year", "")}',
            'Course Code': f'No Open Library  result for {row.get("Course Number", "")}'
        }

        for key, value in row.items():
            if key not in results:
                results[key] = value

    return results
def process_spreadsheet(file_path, api_key_courses, api_bib):
    df = read_excel_file(file_path)
    df = df[(df['EAN-13'] != "") & (~df['EAN-13'].isna())]

    print(len(df))

    print("\n\n\n")
    results = []
    df = df.rename(columns={'Course': 'Numeric'})
    df = df.rename(columns={'Title': 'TITLE'})
    df = df.rename(columns={'Author': 'AUTHOR'})
    pd.set_option('display.max_columns', None)


    df['Course Number'] = str(df['Dept']) + "-" + str(df['Numeric']) + "-" + str(df['Sec'])
    x = 0
    for _, row in df.iterrows():
        # if x == 25:
        #     break
        if x != 0 and x % 100 == 0:
            time.sleep(30)
        x += 1
        result = search_alma_sru_api(row, api_key_courses, api_bib)
        if result:
            results.extend(result)
        else:
            results.append({
                'Title': f'No results for {row.get("TITLE", "")}',
                'Author': f'No results for {row.get("AUTHOR", "")}',
                'Contributor': f'No results format for {row.get("Contributor", "")}',
                'Year': f'No results for {row.get("Year", "")}',
                'Course Code': f'No results {row.get("Course Number", "")}',
                'IBSN': f'No results for {row.get("EAN-13", "")}'
            })

    return results




def generate_excel(results):
    data = []

    def process_result(result):
        # print("before")
        # print(result)

        try:
            results = {
                'Title': result.get('Title', ''),
                'Author': result.get('Author', ''),
                'Contributor': result.get('Contributor', ''),
                'Publisher': result.get('Publisher', ''),
                'Date': result.get('Year', ''),
                'MMS ID': result.get('MMS ID', ''),
                'ISBN': result.get('ISBN', ''),
                'Version': result.get('Version', ''),
                'Course Code': result.get('Course Code', ''),
                'Course Section': result.get('Course Section', ''),
                'Returned Format': result.get('Returned Format', ''),
                'Library': result.get('Library', ''),
                'Location': result.get('Location', ''),
                'Call Number': result.get('Call Number', ''),
                'Barcode': result.get('Barcode', ''),
                'Description': result.get('Description', '')
                #'Citation Type': '',
                #'section_info': result.get('section_info', ''),
                #'Item Policy': result.get('item_policy', result.get('Item Policy', ''))
            }
        except:
            results = {'Title': 'NO RESULTS',
            'Author': 'NO RESULTS',
            'Contributor': 'NO RESULTS',
            'Publisher': r'NO RESULTS',
            'Date': 'NO RESULTS',
            'MMS ID': 'NO RESULTS',
            'ISBN': 'NO RESULTS',
            'Version': 'NO RESULTS',
            'Course Code': 'NO RESULTS',
            'Course Section': 'NO RESULTS',
            'Returned Format': 'NO RESULTS',
            'Library': 'NO RESULTS',
            'Location': 'NO RESULTS',
            'Call Number': 'NO RESULTS',
            'Barcode': 'NO RESULTS',
            'Description': 'NO RESULTS',
            'Citation Type': '',
            'section_info': 'NO RESULTS',
            'Item Policy': 'NO RESULTS'
            }

        # print("after")
        # print(results)
        #print(row['course_code'] + "-" + row['Title'] + "-" + row['MMS ID'])
        # Add any extra fields not already included in the row
        # new_values = {k: v for k, v in results.items() if k not in result}
        # new_row = {**new_values, **result}
        # print("after")
        # print(results)
        # print("\n\n\n")
        return new_row

    # Process each result
    if isinstance(results, list):
        print("list")

        for result in results:
            # print(result)
            # result = { k: ('' if v is None else v) for k, v in result.items() }
            # result = json.loads(result)
            data.append(result)
    else:
        # print("not list")
        # print(results)
        # result = { k: ('' if v is None else v) for k, v in result.items() }
        # result = json.loads(results)
        data.append(results)


    # result = { k: ('' if v is None else v) for k, v in some_dict.items() }
    # Convert to DataFrame
    # print(data)
    result_columns = []
    for k in data[0]:
        # print("data type of list dict 'row'")
        # print(type(data[0]))


        result_columns.append(k)

    # print(result_columns)

    df = pd.DataFrame(columns=result_columns)

    error_file_string = ""
    for list_row in data:
        print(list_row)
        try:

            df = pd.concat([df, pd.DataFrame(list_row, index=[0])])
        except:
            error_file_string = str(list_row) + "\n"

    df = df.reset_index()
    print(df)
    # sys.exit()
    # df = pd.DataFrame(data)

    # print(df)
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Results'

    # Write DataFrame to Excel sheet
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)

            # Handle Barcode formatting
            if df.columns[c_idx] == 'Barcode':
                cell.number_format = numbers.FORMAT_TEXT

    # Set the column headers
    for idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=idx, value=col)

    # Save the workbook to a file
    oDir = "Barnes and Noble Parsed"

    if not os.path.isdir(oDir) or not os.path.exists(oDir):
    	os.makedirs(oDir)

    error_file = open(oDir + '/Barnes and Noble Error File.txt', "w+")

    error_file.write(error_file_string)

    error_file.close()
    wb.save(oDir + '/Barnes and Noble Data with Matching Inventory and Course Codes.xlsx')

def main():
    # Sample results for demonstration
    # results = [
    #     {
    #         'Title': 'Sample Title',
    #         'Author': 'Sample Author',
    #         'Contributor': 'Sample Contributor',
    #         'Publisher': 'Sample Publisher',
    #         'Year': '2024',
    #         'MMS ID': '123456',
    #         'ISBN': '978-3-16-148410-0',
    #         'Version': '1.0',
    #         'course_code': 'CS101',
    #         'course_section': '001',
    #         'Returned Format': 'Physical',
    #         'Library': 'Main Library',
    #         'Location': 'Floor 2',
    #         'Call Number': 'QA76.73.P98 G8 2024',
    #         'Barcode': json.dumps('0123456789'),
    #         'Description': json.dumps('Sample description.'),
    #         'section_info': 'Info Section',
    #         'item_policy': 'Standard'
    #     }
    # ]

    bnf_files = glob.glob('Barnes and Noble/*', recursive = True)

    file_path = bnf_files[0]
    #file_path = askopenfilename(title="Upload Excel file from bookstore")
    api_key_courses = secrets_local.prod_courses_api_key

    api_bib = secrets_local.api_bib


    results = process_spreadsheet(file_path, api_key_courses, api_bib)

    # Generate Excel file
    generate_excel(results)

if __name__ == "__main__":
    main()
